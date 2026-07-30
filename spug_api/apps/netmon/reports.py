# Copyright: (c) OpenSpug Organization. https://github.com/openspug/spug
# Copyright: (c) <spug.dev@gmail.com>
# Released under the AGPL-3.0 License.
"""
报表管理
--------
统计分析：按报表配置的资源范围 + 统计周期，汇总设备可用率、性能均值、异常次数，
生成带图表的 xlsx 报表（复用项目已引入的 openpyxl，无需新增重量级依赖）。
自动发现见 discovery.py；两者结合可满足"资产盘点 + 周期性运行报告"的常规 ITSM 场景。
"""
from django.conf import settings
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from statistics import mean
from libs import human_datetime
from .models import Device, MetricRecord, AnomalyEvent, ReportRecord
from . import stats
import os
import json

HEADER_FILL = PatternFill('solid', fgColor='4F81BD')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')


def _devices_in_scope(report):
    qs = Device.objects.all()
    if report.group_id:
        qs = qs.filter(group_id=report.group_id)
    return qs


def build_report(report, period_start, period_end):
    devices = list(_devices_in_scope(report))
    wb = Workbook()

    # -------- Sheet1: 总览 --------
    ws = wb.active
    ws.title = '总览'
    ws.append([f'{report.name}（{period_start} ~ {period_end}）'])
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(size=14, bold=True)
    ws.append([])
    ws.append(['设备总数', '在线', '告警', '严重', '离线', '异常事件数'])
    _style_header(ws, row=3)

    status_counts = {'online': 0, 'warning': 0, 'critical': 0, 'offline': 0, 'unknown': 0}
    for d in devices:
        status_counts[d.status] = status_counts.get(d.status, 0) + 1
    anomaly_count = AnomalyEvent.objects.filter(
        device__in=devices, created_at__gte=period_start, created_at__lte=period_end
    ).count()
    mttr = stats.compute_mttr(devices, period_start, period_end)
    avail = stats.availability_rate(devices, period_start, period_end)
    ws.append([
        len(devices), status_counts['online'], status_counts['warning'],
        status_counts['critical'], status_counts['offline'], anomaly_count
    ])
    ws.append([])
    ws.append(['整体可用率(%)', avail if avail is not None else '-', 'MTTR平均故障处理时长(分钟)', mttr if mttr is not None else '-'])
    ws['A6'].font = Font(bold=True)
    ws['C6'].font = Font(bold=True)

    chart = BarChart()
    chart.title = '设备状态分布'
    data = Reference(ws, min_col=2, max_col=5, min_row=3, max_row=4)
    cats = Reference(ws, min_col=2, max_col=5, min_row=3, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, 'A8')

    # -------- Sheet2: 设备明细 --------
    ws2 = wb.create_sheet('设备明细')
    ws2.append(['设备名称', 'IP', '分类', '当前状态', '平均CPU(%)', '平均内存(%)', '平均时延(ms)', '异常次数'])
    _style_header(ws2)
    summary_rows = []
    for d in devices:
        metrics_qs = MetricRecord.objects.filter(
            device=d, collected_at__gte=period_start, collected_at__lte=period_end
        )
        cpu_vals = list(metrics_qs.filter(metric_key='cpu').values_list('value', flat=True))
        mem_vals = list(metrics_qs.filter(metric_key='memory').values_list('value', flat=True))
        rtt_vals = list(metrics_qs.filter(metric_key='rtt').values_list('value', flat=True))
        d_anomaly = AnomalyEvent.objects.filter(
            device=d, created_at__gte=period_start, created_at__lte=period_end
        ).count()
        row = [
            d.name, d.ip, d.get_category_display(), d.get_status_display(),
            round(mean(cpu_vals), 2) if cpu_vals else '-',
            round(mean(mem_vals), 2) if mem_vals else '-',
            round(mean(rtt_vals), 2) if rtt_vals else '-',
            d_anomaly,
        ]
        ws2.append(row)
        summary_rows.append(row)

    # -------- Sheet3: 异常事件明细 --------
    ws3 = wb.create_sheet('异常事件')
    ws3.append(['时间', '设备', 'IP', '指标', '当前值', '基线/阈值', '级别', '判定方式', '说明'])
    _style_header(ws3)
    for e in AnomalyEvent.objects.filter(
        device__in=devices, created_at__gte=period_start, created_at__lte=period_end
    ).order_by('-created_at')[:2000]:
        ws3.append([
            e.created_at, e.device.name, e.device.ip, e.metric_key, e.value,
            e.baseline, e.get_level_display(), e.get_method_display(), e.message
        ])

    # -------- Sheet4: TOP故障设备 --------
    ws4 = wb.create_sheet('TOP故障设备')
    ws4.append(['排名', '设备名称', 'IP', '区间内异常次数'])
    _style_header(ws4)
    top_faulty = stats.top_faulty_devices(devices, period_start, period_end, limit=20)
    for i, item in enumerate(top_faulty, 1):
        ws4.append([i, item['device_name'], item['device_ip'], item['count']])
    if top_faulty:
        chart2 = BarChart()
        chart2.title = 'TOP故障设备（异常次数）'
        data = Reference(ws4, min_col=4, max_row=1 + len(top_faulty))
        cats = Reference(ws4, min_col=2, min_row=2, max_row=1 + len(top_faulty))
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        ws4.add_chart(chart2, f'F1')

    for sheet in (ws, ws2, ws3, ws4):
        for column_cells in sheet.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    storage_dir = os.path.join(settings.BASE_DIR, 'storage', 'reports')
    os.makedirs(storage_dir, exist_ok=True)
    filename = f'{report.id}_{human_datetime().replace(":", "").replace(" ", "_").replace("-", "")}.xlsx'
    file_path = os.path.join(storage_dir, filename)
    wb.save(file_path)

    summary = {
        'device_total': len(devices), 'status_counts': status_counts,
        'anomaly_count': anomaly_count, 'mttr_minutes': mttr, 'availability_rate': avail,
        'top_faulty': top_faulty[:5],
    }
    record = ReportRecord.objects.create(
        report=report, period_start=period_start, period_end=period_end,
        file_path=file_path, summary=json.dumps(summary)
    )
    report.last_generated_at = human_datetime()
    report.save(update_fields=['last_generated_at'])
    return record
